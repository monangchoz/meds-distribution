# algorithm_data_export.py
import pandas as pd
import numpy as np
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_cluster_info(filename):
    """
    Extract cluster information from filename
    Returns: number of clusters (1, 3, or 5) or None if not found
    """
    filename_upper = filename.upper()
    if 'NCL_1' in filename_upper:
        return 1
    elif 'NCL_3' in filename_upper:
        return 3
    elif 'NCL_5' in filename_upper:
        return 5
    return None

def extract_data_type(filename):
    """
    Extract data type (historical or generated) from filename
    Returns: 'historical', 'generated', or None if not found
    """
    filename_upper = filename.upper()
    if 'HISTORICAL' in filename_upper:
        return 'historical'
    elif 'GENERATED' in filename_upper:
        return 'generated'
    return None

def load_and_process_csv_files(base_path):
    """
    Load all CSV files from algorithm folders and organize by multiple dimensions:
    - Location (JK2, MKS, SBY)
    - Algorithm (AVNS, GA, DE, BRKGA, PSO)
    - Clusters (1, 3, 5)
    - Data Type (historical, generated)
    """
    # Multi-dimensional dictionary structure
    data_by_location = {
        'JK2': {},
        'MKS': {},
        'SBY': {}
    }
    
    # Also create separate structures for cluster and data type analysis
    data_by_cluster = {1: {}, 3: {}, 5: {}}
    data_by_datatype = {'historical': {}, 'generated': {}}
    
    # Define algorithm folders and their display names
    algorithm_folders = {
        'avns': 'AVNS',
        'ga': 'GA', 
        'de': 'DE',
        'brkga': 'BRKGA',
        'pso': 'PSO'
    }
    
    results_path = os.path.join(base_path, 'results')
    
    if not os.path.exists(results_path):
        print(f"Results folder not found at: {results_path}")
        return data_by_location, data_by_cluster, data_by_datatype
    
    # Process each algorithm folder
    for folder_name, algorithm_name in algorithm_folders.items():
        algorithm_path = os.path.join(results_path, folder_name)
        
        if not os.path.exists(algorithm_path):
            print(f"Algorithm folder not found: {algorithm_path}")
            continue
            
        # Get all CSV files in this algorithm folder
        csv_files = glob.glob(os.path.join(algorithm_path, "*.csv"))
        
        for file_path in csv_files:
            filename = os.path.basename(file_path)
            filename_upper = filename.upper()
            
            # Extract all dimensions
            location = None
            if 'JK2' in filename_upper:
                location = 'JK2'
            elif 'MKS' in filename_upper:
                location = 'MKS'
            elif 'SBY' in filename_upper:
                location = 'SBY'
            
            cluster_count = extract_cluster_info(filename)
            data_type = extract_data_type(filename)
            
            try:
                # Read the CSV file
                df = pd.read_csv(file_path, header=None, names=['total_cost', 'running_time'])
                
                # Store in location-based structure (existing functionality)
                if location:
                    if algorithm_name not in data_by_location[location]:
                        data_by_location[location][algorithm_name] = {'total_cost': [], 'running_time': []}
                    data_by_location[location][algorithm_name]['total_cost'].extend(df['total_cost'].tolist())
                    data_by_location[location][algorithm_name]['running_time'].extend(df['running_time'].tolist())
                
                # Store in cluster-based structure
                if cluster_count is not None:
                    if algorithm_name not in data_by_cluster[cluster_count]:
                        data_by_cluster[cluster_count][algorithm_name] = {'total_cost': [], 'running_time': []}
                    data_by_cluster[cluster_count][algorithm_name]['total_cost'].extend(df['total_cost'].tolist())
                    data_by_cluster[cluster_count][algorithm_name]['running_time'].extend(df['running_time'].tolist())
                
                # Store in data type-based structure
                if data_type:
                    if algorithm_name not in data_by_datatype[data_type]:
                        data_by_datatype[data_type][algorithm_name] = {'total_cost': [], 'running_time': []}
                    data_by_datatype[data_type][algorithm_name]['total_cost'].extend(df['total_cost'].tolist())
                    data_by_datatype[data_type][algorithm_name]['running_time'].extend(df['running_time'].tolist())
                
                # Enhanced logging
                cluster_info = f", Clusters: {cluster_count}" if cluster_count else ""
                datatype_info = f", Data: {data_type}" if data_type else ""
                print(f"Processed {filename} -> Location: {location}, Algorithm: {algorithm_name}{cluster_info}{datatype_info}, Rows: {len(df)}")
                
            except Exception as e:
                print(f"Error processing {filename}: {e}")
    
    return data_by_location, data_by_cluster, data_by_datatype

def create_excel_output(data_by_location, data_by_datatype, output_filename="algorithm_performance_results.xlsx"):
    """
    Create an Excel file with detailed statistics for historical and generated data
    across all locations and algorithms.
    
    Format:
    - Each location gets its own sheet
    - Data is organized by data type (historical/generated)
    - For each algorithm, shows: mean total cost, std dev total cost, mean running time, std dev running time
    """
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Process each location
        locations = ['JK2', 'MKS', 'SBY']
        location_names = ['Jakarta 2', 'Makassar', 'Surabaya']
        
        # First sheet is a summary
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Setup header style
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Create sheets for each location
        for i, (location, location_name) in enumerate(zip(locations, location_names)):
            if i == 0:
                sheet = summary_sheet
            else:
                sheet = wb.create_sheet(location)
            
            # Add title
            sheet['A1'] = f"{location_name} - Algorithm Performance Statistics"
            sheet['A1'].font = Font(bold=True, size=14)
            sheet.merge_cells('A1:G1')
            sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Create header rows
            sheet['A3'] = "Data Type"
            sheet['B3'] = "Algorithm"
            sheet['C3'] = "Mean Total Cost"
            sheet['D3'] = "Std Dev Total Cost"
            sheet['E3'] = "Mean Running Time (s)"
            sheet['F3'] = "Std Dev Running Time (s)"
            
            # Apply header styling
            for cell in sheet['3:3']:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            
            # Collect data for this location
            row = 4
            
            for data_type in ['historical', 'generated']:
                algorithms_processed = []
                
                # Find all algorithms for this location
                all_algorithms = []
                if location in data_by_location:
                    all_algorithms = list(data_by_location[location].keys())
                
                for alg in sorted(all_algorithms):
                    # Check if we have data for this algorithm at this location
                    if location in data_by_location and alg in data_by_location[location]:
                        cost_data = []
                        time_data = []
                        
                        # Find data for this algorithm and location matching the data type
                        if (data_type in data_by_datatype and 
                            alg in data_by_datatype[data_type]):
                            
                            # Get costs and times at this location
                            local_costs = data_by_location[location][alg]['total_cost']
                            local_times = data_by_location[location][alg]['running_time']
                            
                            # Get all costs and times for this data type
                            datatype_costs = data_by_datatype[data_type][alg]['total_cost']
                            datatype_times = data_by_datatype[data_type][alg]['running_time']
                            
                            # Extract costs and times that appear in both sets
                            for i, (cost, time) in enumerate(zip(local_costs, local_times)):
                                # Check if this cost appears in the datatype costs
                                for j, dt_cost in enumerate(datatype_costs):
                                    if abs(cost - dt_cost) < 0.001:  # Floating point comparison
                                        cost_data.append(cost)
                                        time_data.append(time)
                                        break
                        
                        # Only add row if we have data
                        if cost_data:
                            algorithms_processed.append(alg)
                            
                            # Calculate statistics
                            mean_cost = np.mean(cost_data)
                            std_cost = np.std(cost_data)
                            mean_time = np.mean(time_data)
                            std_time = np.std(time_data)
                            
                            # Add data to sheet
                            sheet[f'A{row}'] = data_type.capitalize()
                            sheet[f'B{row}'] = alg
                            sheet[f'C{row}'] = mean_cost
                            sheet[f'D{row}'] = std_cost
                            sheet[f'E{row}'] = mean_time
                            sheet[f'F{row}'] = std_time
                            
                            # Apply styling
                            for col in range(1, 7):
                                cell = sheet.cell(row=row, column=col)
                                cell.border = Side(border_style="thin")
                                if col >= 3:  # Number formatting for statistics
                                    if col in [3, 4]:  # Cost columns
                                        cell.number_format = '#,##0.00'
                                    else:  # Time columns
                                        cell.number_format = '0.0000'
                            
                            row += 1
                
                # Add a blank row between data types if we processed algorithms
                if algorithms_processed:
                    row += 1
            
            # Adjust column widths
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 20
        
        # Save the workbook
        wb.save(output_filename)
        print(f"\n✅ Excel file created successfully: {output_filename}")
        print("Excel file contains:")
        print("  - Sheet for each location (JK2, MKS, SBY)")
        print("  - Organized by data type (Historical, Generated)")
        print("  - Statistics: Mean and Standard Deviation for Total Cost and Running Time")
        
        return True
    
    except Exception as e:
        print(f"\n❌ Error creating Excel file: {e}")
        print("Please make sure you have the required libraries installed:")
        print("  - pandas")
        print("  - numpy")
        print("  - openpyxl")
        return False

def main():
    """
    Main function to run the Excel export
    """
    print("Algorithm Performance Data Excel Export")
    print("=" * 40)
    
    # Get base path
    base_path = os.getcwd()
    print(f"Looking for CSV files in: {base_path}/results/")
    print("Expected folder structure:")
    print("  - results/avns/")
    print("  - results/ga/") 
    print("  - results/de/")
    print("  - results/brkga/")
    print("  - results/pso/")
    print("Expected filename patterns:")
    print("  - Location codes: JK2, MKS, SBY")
    print("  - Data type codes: historical, generated")
    print("-" * 40)
    
    # Load and process data
    data_by_location, data_by_cluster, data_by_datatype = load_and_process_csv_files(base_path)
    
    # Check if data was loaded
    locations_with_data = [loc for loc in ['JK2', 'MKS', 'SBY'] if data_by_location[loc]]
    if not locations_with_data:
        print("❌ No data found! Please check the folder structure and file naming conventions.")
        return
    
    # Get output filename
    default_filename = "algorithm_performance_results.xlsx"
    print("\nEnter filename for Excel output:")
    print(f"(Press Enter to use default: {default_filename})")
    excel_filename = input("> ").strip()
    
    if not excel_filename:
        excel_filename = default_filename
    if not excel_filename.endswith('.xlsx'):
        excel_filename += '.xlsx'
    
    # Create Excel file
    create_excel_output(data_by_location, data_by_datatype, excel_filename)

if __name__ == "__main__":
    main()